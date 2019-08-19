import csv,openpyxl,time,threading,requests,json
import execjs       #用来执行js脚本
from openpyxl.styles import PatternFill,Alignment,Font


n = 0  # 漏洞计数器

host = []   #IP列表
port = []   #端口列表
name = []   #漏洞名称列表
risk = []   #风险等级列表
description = []    #漏洞说明列表
solution = []   #加固建议列表
cve = []    #CVE列表
columns= ['IP','端口','漏洞名称','风险等级','漏洞说明','加固建议','CVE']


class Py4Js():         #js生成tk
 def __init__(self):
    self.ctx = execjs.compile(""" 
    function TL(a) { 
    var k = ""; 
    var b = 406644; 
    var b1 = 3293161072;       
    var jd = "."; 
    var $b = "+-a^+6"; 
    var Zb = "+-3^+b+-f";    
    for (var e = [], f = 0, g = 0; g < a.length; g++) { 
        var m = a.charCodeAt(g); 
        128 > m ? e[f++] = m : (2048 > m ? e[f++] = m >> 6 | 192 : (55296 == (m & 64512) && g + 1 < a.length && 56320 == (a.charCodeAt(g + 1) & 64512) ? (m = 65536 + ((m & 1023) << 10) + (a.charCodeAt(++g) & 1023), 
        e[f++] = m >> 18 | 240, 
        e[f++] = m >> 12 & 63 | 128) : e[f++] = m >> 12 | 224, 
        e[f++] = m >> 6 & 63 | 128), 
        e[f++] = m & 63 | 128) 
    } 
    a = b; 
    for (f = 0; f < e.length; f++) a += e[f], 
    a = RL(a, $b); 
    a = RL(a, Zb); 
    a ^= b1 || 0; 
    0 > a && (a = (a & 2147483647) + 2147483648); 
    a %= 1E6; 
    return a.toString() + jd + (a ^ b) 
  };      
  function RL(a, b) { 
    var t = "a"; 
    var Yb = "+"; 
    for (var c = 0; c < b.length - 2; c += 3) { 
        var d = b.charAt(c + 2), 
        d = d >= t ? d.charCodeAt(0) - 87 : Number(d), 
        d = b.charAt(c + 1) == Yb ? a >>> d: a << d; 
        a = b.charAt(c) == Yb ? a + d & 4294967295 : a ^ d 
    } 
    return a 
  } 
 """)
 def getTk(self,text):
      return self.ctx.call("TL",text)

def buildUrl(text,tk):              #请求包参数
  baseUrl='https://translate.google.cn/translate_a/single'
  baseUrl+='?client=webapp&'
  baseUrl+='sl=en&'
  baseUrl+='tl=zh-CN&'
  baseUrl+='hl=zh-CN&'
  baseUrl+='dt=at&'
  baseUrl+='dt=bd&'
  baseUrl+='dt=ex&'
  baseUrl+='dt=ld&'
  baseUrl+='dt=md&'
  baseUrl+='dt=qca&'
  baseUrl+='dt=rw&'
  baseUrl+='dt=rm&'
  baseUrl+='dt=ss&'
  baseUrl+='dt=t&'
  baseUrl+='source=bh&'
  baseUrl+='ssel=0&'
  baseUrl+='tsel=0&'
  baseUrl+='kc=1&'
  baseUrl+='tk='+str(tk)+'&'
  baseUrl+='q='+text
  return baseUrl

def trans(text):            #谷歌翻译
  header={
    'authority':'translate.google.cn',
    'method':'GET',
    'path':'',
    'scheme':'https',
    'accept':'*/*',
    'accept-encoding':'gzip, deflate, br',
    'accept-language':'zh-CN,zh;q=0.9',
    'cookie':'',
    'user-agent':'Mozilla/5.0 (Windows NT 10.0; WOW64)  AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.108 Safari/537.36',
'x-client-data':'CIa2yQEIpbbJAQjBtskBCPqcygEIqZ3KAQioo8oBGJGjygE='
  }
  url=buildUrl(text,js.getTk(text))
  res = ''
  try:
      r=requests.get(url)
      result=json.loads(r.text)
      if result[7]!=None:
      # 如果我们文本输错，提示你是不是要找xxx的话，那么重新把xxx正确的翻译之后返回
          try:
              correctText=result[7][0].replace('<b><i>',' ').replace('</i></b>','')
              print(correctText)
              correctUrl=buildUrl(correctText,js.getTk(correctText))
              correctR=requests.get(correctUrl)
              newResult=json.loads(correctR.text)
              res=newResult[0][0][0]
          except Exception as e:
              print(e)
              res=result[0][0][0]
      else:
        res=result[0][0][0]
  except Exception as e:
      res=''
      print(url)
      print("翻译"+text+"失败")
      print("错误信息:")
      print(e)
  finally:
      return res
  time.sleep(0.1)


def riskmean(risks):        #风险等级翻译
    if risks == 'Critical':
        m = '紧急'
    elif risks == 'High':
        m = '高危'
    elif risks == 'Medium':
        m = '中危'
    elif risks == 'Low':
        m = '低危'
    risk.append(m)

def ex():                  #创建Excel表，并写入数据
    global n
    wv = openpyxl.Workbook()
    wv.save(filename='Nessus漏洞表.xlsx')
    wb = openpyxl.load_workbook('Nessus漏洞表.xlsx')
    sheet = wb.active
    for i in range(1, len(columns) + 1):  # 写入表头
        sheet.cell(row=1, column=i, value=str(columns[i - 1])).fill = PatternFill("solid", fgColor="CC99FF")
        sheet.cell(row=1, column=i, value=str(columns[i - 1])).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=1, column=i, value=str(columns[i - 1])).font = Font(bold=True)

    for j in range(2,9):
        if j == 2: data = host
        if j == 3: data = port
        if j == 4: data = name
        if j == 5: data = risk
        if j == 6: data = description
        if j == 7: data = solution
        if j == 8: data = cve
        #print(data)
        for i in range(2, len(data) + 1):  # 写入数据
            _ = sheet.cell(row=i, column=j-1, value=str(data[i - 1]))
    sheet.title='扫描结果'
    wb.save('Nessus漏洞表.xlsx')

    Critical = PatternFill("solid", fgColor="FF0000")
    High = PatternFill("solid", fgColor="FFA500")
    Medium = PatternFill("solid", fgColor="FFFF00")
    Low = PatternFill("solid", fgColor="C0FF3E")
    for i in range(2,n+1):
        if sheet.cell(row=i,column=4).value == '紧急':
            sheet.cell(row=i, column=4).fill = Critical
        elif sheet.cell(row=i,column=4).value == '高危':
            sheet.cell(row=i, column=4).fill = High
        elif sheet.cell(row=i,column=4).value == '中危':
            sheet.cell(row=i, column=4).fill = Medium
        elif sheet.cell(row=i,column=4).value == '低危':
            sheet.cell(row=i, column=4).fill = Low
    wb.save('Nessus漏洞表.xlsx')

    print("保存成功,共" + str(n) + '个漏洞')
    print('程序结束：' + time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))

def runs(i):
    global n
    if i[1] != 'CVE':
        if i[3] != 'None':
            host.append(i[4])
            port.append(i[6])
            name.append(trans(i[7]))
            riskmean(i[3])
            description.append(trans(i[9].replace("\n"," ")))
            solution.append(trans(i[10].replace("\n"," ")))
            cve.append(i[1])
            n += 1


if __name__ == '__main__':
    print('程序开始：' + time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))
    js = Py4Js()
    with open("1.csv","r",encoding="utf-8") as csvfile:
        reader = csv.reader(csvfile)
        for i in reader:
            t1 = threading.Thread(target=runs, args=(i,))
            t1.start()
            #time.sleep(0.1)
            t1.join()
    ex()